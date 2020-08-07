import openpyxl

path1=input("enter your downloaded attendance file location: ")
path2=input("enter your class data file location: ")
wb_obj1 = openpyxl.load_workbook(path1)
sheet_obj1 = wb_obj1.active
wb_obj2 = openpyxl.load_workbook(path2)
sheet_obj2 = wb_obj2.active
i=0
alll=[]
attend=[]
no=[]
while(i<sheet_obj2.max_row):
    j=0
    
    cell_obj2 = sheet_obj2.cell(row = i+2, column = 1)
    if(cell_obj2.value==None):
        break
    rno = sheet_obj2.cell(row = i+2, column = 2)
    alll.append(cell_obj2.value)
    no.append(rno.value)
    #print(cell_obj2.value)
    while(j<sheet_obj1.max_row):
        cell_obj1 = sheet_obj1.cell(row = j+2, column = 1)
        if(cell_obj1.value==None):
            break
        j=j+1
        if(cell_obj1.value==cell_obj2.value):
            attend.append(cell_obj2.value)

    i=i+1
a = [] 
for num in attend: 
    if num not in a: 
        a.append(num)

def new_file():
    path5=input("enter the name of the file with its path: ")
    neww = openpyxl.Workbook()
    new_sheet=neww.active
    i=0
    c1 = new_sheet.cell(row = 1, column = 1)
    c1.value = "Name"
    c1 = new_sheet.cell(row = 1, column = 2)
    c1.value = "RollNo"
    c1 = new_sheet.cell(row = 1, column = 3)
    date=input("Enter a date: ")
    c1.value =  date
    c1 = new_sheet.cell(row = 2, column = 3)
    c1.value=input("Enter the hour/subject name: ")
    c1 = new_sheet.cell(row = 3, column = 3)
    c1.value=str(1)
    while(i<len(alll)):
        j=0
        flag=0
        while(j<len(attend)):
            if(alll[i]==attend[j]):
                c1 = new_sheet.cell(row = i+4, column = 1)
                c1.value = alll[i]
                c1 = new_sheet.cell(row = i+4, column = 2)
                c1.value = no[i]
                c1 = new_sheet.cell(row = i+4, column = 3)
                c1.value = "1"
                c1.value=str(c1.value)
                flag=1
            j=j+1
        if(flag==0):
            c1 = new_sheet.cell(row = i+4, column = 1)
            c1.value = alll[i]
            c1 = new_sheet.cell(row = i+4, column = 2)
            c1.value = no[i]
            c1 = new_sheet.cell(row = i+4, column = 3)
            c1.value = "ab|0"
            c1.value=str(c1.value)
        i=i+1
    c1=new_sheet.cell(row = i+5, column = 3)
    c1.value = len(a)
    c1=new_sheet.cell(row = i+6, column = 3)
    c1.value = len(alll)-len(a)
    neww.save(path5)

def existing_file():
    path3=input("enter your excel sheet attendence file location: ")
    neww = openpyxl.load_workbook(path3)
    new_sheet = neww.active
    col=new_sheet.max_column
    i=0
    c1 = new_sheet.cell(row = 1, column = col+1)
    date=input("Enter a date: ")
    c1.value =  date
    c1 = new_sheet.cell(row = 2, column = col+1)
    c1.value=input("Enter the hour/subject name: ")
    c1 = new_sheet.cell(row = 3, column = col+1)
    c2 = new_sheet.cell(row = 3, column = col)
    c1.value=int(c2.value)+1
    c1.value=str(c1.value)
    while(i<len(alll)):
        j=0
        flag=0
        while(j<len(attend)):
            if(alll[i]==attend[j]):
                c1 = new_sheet.cell(row = i+4, column = col+1)
                c2 = new_sheet.cell(row = i+4, column = col)
                if(c2.value[:3]=="ab|"):
                    c1.value=int(c2.value.replace('ab|',''))+1
                    c1.value=str(c1.value)
                else:
                    c1.value = int(c2.value)+1
                    c1.value=str(c1.value)
                flag=1
            j=j+1
        if(flag==0):
            c1 = new_sheet.cell(row = i+4, column = col+1)
            c2 = new_sheet.cell(row = i+4, column = col)
            c2.value=str(c2.value)
            if(c2.value[:3]=="ab|"):
                c1.value=c2.value
                c1.value=str(c1.value)
            else:
                c1.value = "ab|"+str(c2.value)
                c1.value=str(c1.value)
            
        i=i+1
    c1=new_sheet.cell(row = i+5, column = col+1)
    c1.value = len(a)
    c1=new_sheet.cell(row = i+6, column = col+1)
    c1.value =len(alll)-len(a)
    filen=path3
    neww.save(filen)

def cal_per():
    path4=input("enter your excel sheet attendence file location: ")
    neww = openpyxl.load_workbook(path4)
    new_sheet = neww.active
    col=new_sheet.max_column
    i=0
    c1 = new_sheet.cell(row = 1, column = col+2)
    c1.value="Total Absent"
    c1 = new_sheet.cell(row = 1, column = col+1)
    c1.value="Total Present"
    c3 = new_sheet.cell(row = 3, column = col)
    c5 = new_sheet.cell(row = 1, column = col+3)
    c5.value="Percentage"
    while(i<len(alll)):
        j=0
        flag=0
        while(j<len(attend)):
            if(alll[i]==attend[j]):
                c1 = new_sheet.cell(row = i+4, column = col+1)
                c2 = new_sheet.cell(row = i+4, column = col)
                c4 = new_sheet.cell(row = i+4, column = col+2)
                c5 = new_sheet.cell(row = i+4, column = col+3)
                if(c2.value[:3]=="ab|"):
                    c1.value=int(c2.value.replace('ab|',''))
                    c4.value=int(c3.value)-c1.value
                    c5.value=(float(c1.value)*100)/float(c3.value)
                    c1.value=str(c1.value)
                    c4.value=str(c4.value)
                    c5.value=str(c5.value)
                else:
                    c1.value = int(c2.value)
                    c4.value=int(c3.value)-c1.value
                    c5.value=(float(c1.value)*100)/float(c3.value)
                    c1.value=str(c1.value)
                    c4.value=str(c4.value)
                    c5.value=str(c5.value)
                flag=1
            j=j+1
        if(flag==0):
            c1 = new_sheet.cell(row = i+4, column = col+1)
            c2 = new_sheet.cell(row = i+4, column = col)
            c4 = new_sheet.cell(row = i+4, column = col+2)
            c5 = new_sheet.cell(row = i+4, column = col+3)
            c2.value=str(c2.value)
            if(c2.value[:3]=="ab|"):
                c1.value=c2.value
                c4.value=int(c3.value)-int(c1.value[3:])
                c5.value=(float(c1.value[3:])*100)/float(c3.value)
                c1.value=str(c1.value.replace('ab|',''))
                c4.value=str(c4.value)
                c5.value=str(c5.value)
            else:
                c4.value=int(c3.value)-int(c1.value[3:])
                c5.value=(float(c1.value[3:])*100)/float(c3.value)
                c1.value = str(c2.value.replace('ab|',''))
                c4.value=str(c4.value)
                c5.value=str(c5.value)
        i=i+1


    neww.save(path4)

    
print("1. New file")
print("2. Existing file")
print("3. Calculate percentage")
x=int(input("select your choice: "))
if(x==1):
    new_file()
elif(x==2):
    existing_file()
elif(x==3):
    cal_per()
#r"F:\try\try.xlsx"
#r"F:\try\data1.xlsx"
#r"F:\try\new.xlsx"
    
